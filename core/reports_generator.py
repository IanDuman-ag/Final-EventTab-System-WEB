import io
from datetime import datetime
from django.contrib.admin.models import LogEntry, DELETION, CHANGE, ADDITION
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ReportLab imports for PDF generation
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

from events.models import (
    Event, BracketTeam, BracketMatch, ScoreSheet,
    JudgingEvent, Candidate, Criterion, JudgeScore,
)

class NumberedCanvas(canvas.Canvas):
    """
    Two-pass canvas to dynamically compute and display 'Page X of Y' in PDF footer.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        
        # Primary Color Palette (Navy: #021943, Accents: #ffdf25)
        navy = colors.HexColor('#021943')
        grey = colors.HexColor('#6d8098')
        light_grey = colors.HexColor('#e6edf7')

        # Header (Top Bar)
        self.setStrokeColor(light_grey)
        self.setLineWidth(1)
        self.line(36, 756, 576, 756)
        
        self.setFont("Helvetica-Bold", 8)
        self.setFillColor(navy)
        self.drawString(36, 762, "EVENTTAB MANAGEMENT SYSTEM")
        
        self.setFont("Helvetica", 8)
        self.setFillColor(grey)
        self.drawRightString(576, 762, datetime.now().strftime("%B %d, %Y - %I:%M %p"))

        # Footer (Bottom Bar)
        self.line(36, 48, 576, 48)
        self.drawString(36, 36, "CONFIDENTIAL - SUPER ADMIN REPORT EXPORT")
        
        page_str = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(576, 36, page_str)
        self.restoreState()


TOURNAMENT_TYPE_LABELS = {
    'single_elimination': 'Single Elimination',
    'double_elimination': 'Double Elimination',
    'round_robin': 'Round Robin',
    'best_of_3': 'Best of 3',
    'best_of_5': 'Best of 5',
}


def _derive_event_type(event_obj):
    """Return ('match'|'criteria', 'Match Result Based'|'Criteria Based')."""
    method = (getattr(event_obj, 'scoring_method', '') or '').strip().lower()
    if method in ('match', 'criteria'):
        return method, 'Match Result Based' if method == 'match' else 'Criteria Based'
    # Fall back to category heuristic (mirrors Manage Events page)
    cat = (getattr(event_obj, 'category', '') or '').strip().lower()
    if cat in ('sports', 'esports', 'e-sports'):
        return 'match', 'Match Result Based'
    return 'criteria', 'Criteria Based'


def _build_bracket_overview(event_obj, round_filter=None, status_filter=None):
    """Group BracketMatch records by round into a tournament-tree structure."""
    matches_qs = (
        BracketMatch.objects.filter(event=event_obj)
        .select_related('team_a', 'team_b', 'winner')
        .order_by('match_number')
    )
    if round_filter:
        matches_qs = matches_qs.filter(round_name__icontains=round_filter)
    if status_filter:
        matches_qs = matches_qs.filter(status__iexact=status_filter)

    rounds = {}
    final_winner = 'TBD'
    for m in matches_qs:
        rounds.setdefault(m.round_name or 'Round', []).append({
            'match_no': m.match_number,
            'team_a': m.team_a.name if m.team_a else 'TBD',
            'team_b': m.team_b.name if m.team_b else 'TBD',
            'score_a': m.score_a or '—',
            'score_b': m.score_b or '—',
            'winner': m.winner.name if m.winner else 'TBD',
            'status': m.get_status_display(),
        })
        if m.winner and (m.round_name or '').lower().find('final') != -1:
            final_winner = m.winner.name

    bracket_rounds = [{'round': rname, 'matches': mlist} for rname, mlist in rounds.items()]
    return bracket_rounds, final_winner


def _build_criteria_results(event_obj):
    """Per-candidate criterion scores, weighted total, and rank (criteria-based events)."""
    je = getattr(event_obj, 'judging_event', None)
    if not je:
        return [], []

    criteria = list(Criterion.objects.filter(event=je).order_by('order'))
    candidates = list(Candidate.objects.filter(event=je).order_by('number'))
    if not criteria or not candidates:
        return [], []

    crit_names = [c.name for c in criteria]
    results = []
    for cand in candidates:
        per_criterion = {}
        weighted_total = 0.0
        for crit in criteria:
            scores = list(
                JudgeScore.objects.filter(candidate=cand, criterion=crit)
                .values_list('score', flat=True)
            )
            avg = (sum(float(s) for s in scores) / len(scores)) if scores else 0.0
            per_criterion[crit.name] = round(avg, 2)
            weighted_total += avg * (float(crit.weight_percent or 0) / 100.0)
        results.append({
            'candidate': cand.name,
            'number': cand.number,
            'per_criterion': per_criterion,
            'total': round(weighted_total, 2),
        })

    results.sort(key=lambda r: r['total'], reverse=True)
    for idx, r in enumerate(results, 1):
        r['rank'] = idx
    return results, crit_names


def _build_leaderboard(event_obj, event_type, criteria_results):
    """Standings/points summary. Match → wins; Criteria → weighted totals."""
    leaderboard = []
    if event_type == 'criteria':
        for r in criteria_results:
            leaderboard.append({
                'rank': r['rank'],
                'name': r['candidate'],
                'group': f"#{r['number']}",
                'points': r['total'],
                'detail': 'Weighted score',
            })
        return leaderboard

    # Match-based: rank teams by wins (then fewest losses)
    teams = list(BracketTeam.objects.filter(event=event_obj).select_related('department'))
    standings = []
    for t in teams:
        wins = BracketMatch.objects.filter(event=event_obj, winner=t).count()
        standings.append({
            'name': t.name,
            'group': t.department.name if t.department else (t.members[:30] if t.members else '—'),
            'wins': wins,
            'losses': t.loss_count,
        })
    standings.sort(key=lambda s: (s['wins'], -s['losses']), reverse=True)
    for idx, s in enumerate(standings, 1):
        leaderboard.append({
            'rank': idx,
            'name': s['name'],
            'group': s['group'],
            'points': s['wins'],
            'detail': f"{s['wins']}W - {s['losses']}L",
        })
    return leaderboard


def get_reporting_data(event_name, department_filter="all", start_date=None, end_date=None,
                       round_filter=None, status_filter=None, category_filter=None,
                       division_filter=None):
    """
    Gathers dynamic reporting data from the database and returns a single dict.
    If no events/scores exist, falls back to high-fidelity mock data.

    Returned keys: event_details, rankings, participants, audit_trail,
    bracket_rounds, final_winner, criteria_results, criteria_names, leaderboard.
    """
    event_obj = Event.objects.filter(name=event_name).first()

    if event_obj:
        event_type_key, event_type_label = _derive_event_type(event_obj)
        tournament_label = TOURNAMENT_TYPE_LABELS.get(
            (event_obj.tournament_type or '').strip().lower(),
            event_obj.tournament_type or 'N/A',
        )
        assigned_judges = ', '.join(
            (f"{j.first_name} {j.last_name}".strip() or j.username)
            for j in event_obj.assigned_judges.all()
        ) or 'None assigned'

        event_details = {
            'name': event_obj.name,
            'category': event_obj.category,
            'division': event_obj.division or 'Unassigned',
            'department': event_obj.department or 'Unassigned',
            'date': event_obj.event_date.strftime('%B %d, %Y') if event_obj.event_date else 'TBD',
            'time': event_obj.event_time.strftime('%I:%M %p') if event_obj.event_time else 'TBD',
            'venue': event_obj.venue,
            'status': event_obj.get_status_display(),
            'max_participants': event_obj.max_participants or 'Unlimited',
            'total_teams': event_obj.num_teams or event_obj.bracket_teams.count(),
            'event_type': event_type_label,
            'tournament_type': tournament_label,
            'assigned_judges': assigned_judges,
        }

        # Match results (from ScoreSheet)
        score_sheets = ScoreSheet.objects.filter(event=event_obj).select_related('match', 'winner', 'tabulator')
        if status_filter:
            score_sheets = score_sheets.filter(status__iexact=status_filter)
        rankings = []
        for index, ss in enumerate(score_sheets, 1):
            rankings.append({
                'rank': index,
                'team_a': ss.match.team_a.name if ss.match.team_a else 'TBD',
                'team_b': ss.match.team_b.name if ss.match.team_b else 'TBD',
                'score_a': float(ss.score_team_a),
                'score_b': float(ss.score_team_b),
                'winner': ss.winner.name if ss.winner else 'TBD',
                'match_no': ss.match.match_number,
                'status': ss.get_status_display(),
                'tabulator': ss.tabulator.username if ss.tabulator else 'System',
                'remarks': ss.remarks or 'N/A'
            })

        # Participants / Teams
        teams_qs = BracketTeam.objects.filter(event=event_obj).select_related('department')
        if department_filter and department_filter != 'all':
            teams_qs = teams_qs.filter(department__name=department_filter)
        participants = []
        for t in teams_qs:
            participants.append({
                'seed': t.seed,
                'name': t.name,
                'department': t.department.name if t.department else 'Unassigned',
                'loss_count': t.loss_count,
                'members': t.members or 'No members declared'
            })

        # Bracket overview, criteria results, leaderboard
        bracket_rounds, final_winner = _build_bracket_overview(event_obj, round_filter, status_filter)
        criteria_results, criteria_names = _build_criteria_results(event_obj)
        leaderboard = _build_leaderboard(event_obj, event_type_key, criteria_results)

        # Audit Trail (LogEntry)
        logs_qs = LogEntry.objects.filter(object_id=str(event_obj.id)).select_related('user').order_by('-action_time')
        if start_date:
            logs_qs = logs_qs.filter(action_time__date__gte=start_date)
        if end_date:
            logs_qs = logs_qs.filter(action_time__date__lte=end_date)
        audit_trail = []
        for log in logs_qs:
            action_type = "Created" if log.is_addition() else "Modified" if log.is_change() else "Deleted"
            audit_trail.append({
                'timestamp': log.action_time.strftime('%Y-%m-%d %I:%M %p'),
                'user': log.user.username if log.user else 'System',
                'action_type': action_type,
                'repr': log.object_repr,
                'message': log.change_message or 'No details provided'
            })

    else:
        # High-fidelity mock data fallback (when Event doesn't exist yet)
        event_details = {
            'name': event_name or 'National Science Decathlon 2024',
            'category': 'Academic Competition',
            'division': 'Seniors',
            'department': 'College of Science',
            'date': 'October 24, 2024',
            'time': '09:00 AM',
            'venue': 'Main Assembly Hall',
            'status': 'Active',
            'max_participants': 200,
            'total_teams': 8,
            'event_type': 'Match Result Based',
            'tournament_type': 'Single Elimination',
            'assigned_judges': 'Judge Angela, Judge Brandon, Judge Carter',
        }

        rankings = [
            {'rank': 1, 'team_a': 'Quantum Physicists', 'team_b': 'Bio-Warriors', 'score_a': 94.50, 'score_b': 88.00, 'winner': 'Quantum Physicists', 'match_no': 1, 'status': 'Finalized', 'tabulator': 'adminn', 'remarks': 'Outstanding performance in theory rounds.'},
            {'rank': 2, 'team_a': 'Alchemists', 'team_b': 'Neuro-Hackers', 'score_a': 91.00, 'score_b': 92.50, 'winner': 'Neuro-Hackers', 'match_no': 2, 'status': 'Finalized', 'tabulator': 'adminn', 'remarks': 'Extremely close matching in final minutes.'},
            {'rank': 3, 'team_a': 'Astrophysicists', 'team_b': 'Geodesics', 'score_a': 85.00, 'score_b': 81.50, 'winner': 'Astrophysicists', 'match_no': 3, 'status': 'Finalized', 'tabulator': 'tabulator', 'remarks': 'Advancing to quarter-finals.'},
            {'rank': 4, 'team_a': 'Ecology Champions', 'team_b': 'Cybernetics', 'score_a': 76.50, 'score_b': 89.00, 'winner': 'Cybernetics', 'match_no': 4, 'status': 'Finalized', 'tabulator': 'tabulator', 'remarks': 'Scored high on presentation.'},
        ]

        participants = [
            {'seed': 1, 'name': 'Quantum Physicists', 'department': 'Physics Department', 'loss_count': 0, 'members': 'Alice Smith, Bob Jones, Charles Lee'},
            {'seed': 2, 'name': 'Neuro-Hackers', 'department': 'Biology Department', 'loss_count': 0, 'members': 'David Miller, Emma Davis, Frank Wilson'},
            {'seed': 3, 'name': 'Alchemists', 'department': 'Chemistry Department', 'loss_count': 1, 'members': 'Grace Taylor, Henry Harris, Irene Clark'},
            {'seed': 4, 'name': 'Cybernetics', 'department': 'Computer Science Department', 'loss_count': 0, 'members': 'Jack Lewis, Karen Walker, Leo Young'},
        ]

        bracket_rounds = [
            {'round': 'Semifinals', 'matches': [
                {'match_no': 1, 'team_a': 'Quantum Physicists', 'team_b': 'Bio-Warriors', 'score_a': '94', 'score_b': '88', 'winner': 'Quantum Physicists', 'status': 'Completed'},
                {'match_no': 2, 'team_a': 'Neuro-Hackers', 'team_b': 'Astrophysicists', 'score_a': '92', 'score_b': '85', 'winner': 'Neuro-Hackers', 'status': 'Completed'},
            ]},
            {'round': 'Finals', 'matches': [
                {'match_no': 3, 'team_a': 'Quantum Physicists', 'team_b': 'Neuro-Hackers', 'score_a': '96', 'score_b': '90', 'winner': 'Quantum Physicists', 'status': 'Completed'},
            ]},
        ]
        final_winner = 'Quantum Physicists'

        criteria_results = []
        criteria_names = []

        leaderboard = [
            {'rank': 1, 'name': 'Quantum Physicists', 'group': 'Physics Department', 'points': 2, 'detail': '2W - 0L'},
            {'rank': 2, 'name': 'Neuro-Hackers', 'group': 'Biology Department', 'points': 1, 'detail': '1W - 1L'},
            {'rank': 3, 'name': 'Bio-Warriors', 'group': 'Biology Department', 'points': 0, 'detail': '0W - 1L'},
        ]

        audit_trail = [
            {'timestamp': '2026-05-17 10:15 AM', 'user': 'superadmin', 'action_type': 'Created', 'repr': 'National Science Decathlon 2024', 'message': 'Registered decathlon tournament event.'},
            {'timestamp': '2026-05-17 10:30 AM', 'user': 'superadmin', 'action_type': 'Modified', 'repr': 'National Science Decathlon 2024', 'message': 'Generated single-elimination bracket matchups.'},
            {'timestamp': '2026-05-17 11:45 AM', 'user': 'adminn', 'action_type': 'Modified', 'repr': 'Scoresheet Match 1', 'message': 'Submitted initial round scoresheet.'},
            {'timestamp': '2026-05-17 12:00 PM', 'user': 'superadmin', 'action_type': 'Modified', 'repr': 'Scoresheet Match 1', 'message': 'Approved scoresheet and advanced Quantum Physicists.'},
        ]

        if department_filter and department_filter != 'all':
            participants = [p for p in participants if p['department'].lower() == department_filter.lower()]

    return {
        'event_details': event_details,
        'rankings': rankings,
        'participants': participants,
        'audit_trail': audit_trail,
        'bracket_rounds': bracket_rounds,
        'final_winner': final_winner,
        'criteria_results': criteria_results,
        'criteria_names': criteria_names,
        'leaderboard': leaderboard,
    }


def generate_excel_report(event_name, department_filter="all", start_date=None, end_date=None, included_metrics=None, event_info=None):
    """
    Creates an Excel spreadsheet in memory using openpyxl, applying EventTab styling.
    """
    if included_metrics is None:
        included_metrics = ['rankings', 'participants', 'scores']

    event_info = event_info or {}
    data = get_reporting_data(
        event_name, department_filter, start_date, end_date,
        round_filter=event_info.get('round_filter'),
        status_filter=event_info.get('status_filter'),
        category_filter=event_info.get('category_filter'),
        division_filter=event_info.get('division_filter'),
    )
    event_details = data['event_details']
    rankings = data['rankings']
    participants = data['participants']
    audit_trail = data['audit_trail']
    bracket_rounds = data['bracket_rounds']
    final_winner = data['final_winner']
    criteria_results = data['criteria_results']
    criteria_names = data['criteria_names']
    leaderboard = data['leaderboard']

    wb = Workbook()
    
    # ------------------ STYLING CONFIGS ------------------
    font_family = "Segoe UI"
    
    # Styles & Colors
    header_fill = PatternFill(start_color="021943", end_color="021943", fill_type="solid")
    stripe_fill = PatternFill(start_color="F5F8FC", end_color="F5F8FC", fill_type="solid")
    accent_fill = PatternFill(start_color="FFFDF0", end_color="FFFDF0", fill_type="solid")
    
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=11, bold=True, color="021943")
    title_font = Font(name=font_family, size=16, bold=True, color="021943")
    regular_font = Font(name=font_family, size=10, color="333333")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    border_thin = Side(border_style="thin", color="D1DCEF")
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    # 1. SUMMARY SHEET
    ws_summary = wb.active
    ws_summary.title = "Event Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    ws_summary["A1"] = "EVENT SUMMARY REPORT"
    ws_summary["A1"].font = title_font
    ws_summary.merge_cells("A1:C1")
    ws_summary.row_dimensions[1].height = 30
    
    ws_summary["A3"] = "Metric"
    ws_summary["A3"].font = header_font
    ws_summary["A3"].fill = header_fill
    ws_summary["A3"].alignment = align_left
    
    ws_summary["B3"] = "Value"
    ws_summary["B3"].font = header_font
    ws_summary["B3"].fill = header_fill
    ws_summary["B3"].alignment = align_left
    ws_summary.merge_cells("B3:C3")
    ws_summary.row_dimensions[3].height = 24
    
    summary_rows = [
        ("Event Name", event_details['name']),
        ("Event Type", event_details.get('event_type', 'N/A')),
        ("Category", event_details['category']),
        ("Division", event_details['division']),
        ("Department", event_details['department']),
        ("Tournament Type", event_details.get('tournament_type', 'N/A')),
        ("Date", event_details['date']),
        ("Time", event_details['time']),
        ("Venue", event_details['venue']),
        ("Status", event_details['status']),
        ("Assigned Judges / Scorers", event_details.get('assigned_judges', 'None assigned')),
        ("Max Participants", event_details['max_participants']),
        ("Total Teams Registered", event_details['total_teams']),
        ("Generated By", event_info.get('generated_by', 'Admin')),
        ("Generated On", datetime.now().strftime('%B %d, %Y - %I:%M %p')),
        ("Remarks", event_info.get('remarks', 'N/A')),
    ]
    
    for idx, (label, val) in enumerate(summary_rows, 4):
        ws_summary.row_dimensions[idx].height = 20
        ws_summary.cell(row=idx, column=1, value=label).font = bold_font
        ws_summary.cell(row=idx, column=1).border = cell_border
        
        ws_summary.cell(row=idx, column=2, value=val).font = regular_font
        ws_summary.cell(row=idx, column=2).border = cell_border
        ws_summary.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=3)
        ws_summary.cell(row=idx, column=3).border = cell_border

    # 2. RANKINGS & SCORING SHEET
    if 'rankings' in included_metrics or 'scores' in included_metrics:
        ws_rank = wb.create_sheet(title="Rankings & Scores")
        ws_rank.views.sheetView[0].showGridLines = True
        
        ws_rank["A1"] = "RANKINGS & SCORECARDS"
        ws_rank["A1"].font = title_font
        ws_rank.merge_cells("A1:H1")
        ws_rank.row_dimensions[1].height = 30
        
        headers = ["Rank", "Team A", "Team B", "Score A", "Score B", "Match Winner", "Match #", "Status"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws_rank.cell(row=3, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center if col_idx in [1, 4, 5, 7, 8] else align_left
            cell.border = cell_border
        ws_rank.row_dimensions[3].height = 24
        
        for r_idx, r in enumerate(rankings, 4):
            ws_rank.row_dimensions[r_idx].height = 20
            row_vals = [r['rank'], r['team_a'], r['team_b'], r['score_a'], r['score_b'], r['winner'], r['match_no'], r['status']]
            
            for c_idx, val in enumerate(row_vals, 1):
                cell = ws_rank.cell(row=r_idx, column=c_idx, value=val)
                cell.font = regular_font
                cell.border = cell_border
                cell.alignment = align_center if c_idx in [1, 4, 5, 7, 8] else align_left
                
                # Stripe alternate rows
                if r_idx % 2 == 0:
                    cell.fill = stripe_fill

    # 3. PARTICIPANTS SHEET
    if 'participants' in included_metrics:
        ws_part = wb.create_sheet(title="Participants")
        ws_part.views.sheetView[0].showGridLines = True
        
        ws_part["A1"] = "PARTICIPANTS DIRECTORY"
        ws_part["A1"].font = title_font
        ws_part.merge_cells("A1:E1")
        ws_part.row_dimensions[1].height = 30
        
        headers = ["Seed", "Team Name", "Department / Course", "Losses", "Registered Members"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws_part.cell(row=3, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center if col_idx in [1, 4] else align_left
            cell.border = cell_border
        ws_part.row_dimensions[3].height = 24
        
        for r_idx, p in enumerate(participants, 4):
            ws_part.row_dimensions[r_idx].height = 20
            row_vals = [p['seed'], p['name'], p['department'], p['loss_count'], p['members']]
            
            for c_idx, val in enumerate(row_vals, 1):
                cell = ws_part.cell(row=r_idx, column=c_idx, value=val)
                cell.font = regular_font
                cell.border = cell_border
                cell.alignment = align_center if c_idx in [1, 4] else align_left
                
                if r_idx % 2 == 0:
                    cell.fill = stripe_fill

    # ── Helper to write a simple titled table sheet ──
    def _write_table_sheet(title, sheet_title, headers, rows, center_cols=()):
        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True
        last_col = get_column_letter(max(len(headers), 1))
        ws["A1"] = title
        ws["A1"].font = title_font
        ws.merge_cells(f"A1:{last_col}1")
        ws.row_dimensions[1].height = 30
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center if col_idx in center_cols else align_left
            cell.border = cell_border
        ws.row_dimensions[3].height = 24
        for r_idx, row_vals in enumerate(rows, 4):
            ws.row_dimensions[r_idx].height = 20
            for c_idx, val in enumerate(row_vals, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = regular_font
                cell.border = cell_border
                cell.alignment = align_center if c_idx in center_cols else align_left
                if r_idx % 2 == 0:
                    cell.fill = stripe_fill
        return ws

    # 4. BRACKET OVERVIEW SHEET
    if bracket_rounds:
        bracket_rows = []
        for rnd in bracket_rounds:
            for m in rnd['matches']:
                bracket_rows.append([
                    rnd['round'], m['match_no'],
                    f"{m['team_a']} vs {m['team_b']}",
                    f"{m['score_a']} - {m['score_b']}",
                    m['winner'], m['status'],
                ])
        bracket_rows.append([])
        bracket_rows.append(['CHAMPION / FINAL WINNER', '', '', '', final_winner, ''])
        _write_table_sheet(
            "BRACKET OVERVIEW", "Bracket Overview",
            ["Round", "Match #", "Matchup", "Score", "Winner", "Status"],
            bracket_rows, center_cols=(2, 4, 6),
        )

    # 5. CRITERIA RESULTS SHEET
    if criteria_results:
        crit_headers = ["Rank", "Candidate", "No."] + criteria_names + ["Total Score"]
        crit_rows = []
        for r in criteria_results:
            row = [r['rank'], r['candidate'], r['number']]
            row += [r['per_criterion'].get(cn, 0) for cn in criteria_names]
            row.append(r['total'])
            crit_rows.append(row)
        center = tuple([1, 3] + list(range(4, 4 + len(criteria_names) + 1)))
        _write_table_sheet(
            "CRITERIA-BASED RESULTS", "Criteria Results",
            crit_headers, crit_rows, center_cols=center,
        )

    # 6. LEADERBOARD SHEET
    if leaderboard:
        lb_rows = [[l['rank'], l['name'], l['group'], l['points'], l['detail']] for l in leaderboard]
        _write_table_sheet(
            "LEADERBOARD / POINTS SUMMARY", "Leaderboard",
            ["Rank", "Team / Participant", "Group", "Points", "Record / Detail"],
            lb_rows, center_cols=(1, 4),
        )

    # 7. AUDIT TRAIL SHEET
    ws_audit = wb.create_sheet(title="Audit Logs")
    ws_audit.views.sheetView[0].showGridLines = True
    
    ws_audit["A1"] = "SYSTEM AUDIT TRAIL"
    ws_audit["A1"].font = title_font
    ws_audit.merge_cells("A1:E1")
    ws_audit.row_dimensions[1].height = 30
    
    headers = ["Timestamp", "Performed By", "Action Flag", "Component Affected", "Details"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_audit.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center if col_idx in [1, 3] else align_left
        cell.border = cell_border
    ws_audit.row_dimensions[3].height = 24
    
    for r_idx, l in enumerate(audit_trail, 4):
        ws_audit.row_dimensions[r_idx].height = 20
        row_vals = [l['timestamp'], l['user'], l['action_type'], l['repr'], l['message']]
        
        for c_idx, val in enumerate(row_vals, 1):
            cell = ws_audit.cell(row=r_idx, column=c_idx, value=val)
            cell.font = regular_font
            cell.border = cell_border
            cell.alignment = align_center if c_idx in [1, 3] else align_left
            
            if r_idx % 2 == 0:
                cell.fill = stripe_fill

    # Autofit Column Widths across all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            # Avoid setting huge column widths for merged header cell A1
            max_len = 0
            for cell in col:
                if cell.row != 1 and cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Save to dynamic in-memory stream
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream


def generate_pdf_report(event_name, department_filter="all", start_date=None, end_date=None, included_metrics=None, event_info=None):
    """
    Creates a styled PDF report using reportlab, conforming to the EventTab design system.
    """
    if included_metrics is None:
        included_metrics = ['rankings', 'participants', 'scores']

    event_info = event_info or {}
    data = get_reporting_data(
        event_name, department_filter, start_date, end_date,
        round_filter=event_info.get('round_filter'),
        status_filter=event_info.get('status_filter'),
        category_filter=event_info.get('category_filter'),
        division_filter=event_info.get('division_filter'),
    )
    event_details = data['event_details']
    rankings = data['rankings']
    participants = data['participants']
    audit_trail = data['audit_trail']
    bracket_rounds = data['bracket_rounds']
    final_winner = data['final_winner']
    criteria_results = data['criteria_results']
    criteria_names = data['criteria_names']
    leaderboard = data['leaderboard']

    pdf_stream = io.BytesIO()
    
    # 0.5 inches margin (36pt)
    doc = SimpleDocTemplate(
        pdf_stream, 
        pagesize=letter, 
        leftMargin=36, 
        rightMargin=36, 
        topMargin=54, 
        bottomMargin=54
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Brand Colors
    navy = colors.HexColor('#021943')
    yellow = colors.HexColor('#ffdf25')
    text_color = colors.HexColor('#1a2d42')
    border_color = colors.HexColor('#d0dcf0')
    stripe_color = colors.HexColor('#f5f8fc')
    
    # Modify default styles or declare custom ones
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=navy,
        spaceAfter=6
    )
    
    subtitle_style = ParagraphStyle(
        'DocSub',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#6d8098'),
        spaceAfter=18
    )

    h2_style = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=navy,
        spaceBefore=14,
        spaceAfter=8
    )
    
    body_bold = ParagraphStyle(
        'BodyBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=text_color
    )

    body_regular = ParagraphStyle(
        'BodyReg',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=text_color
    )

    body_center = ParagraphStyle(
        'BodyCenter',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=text_color,
        alignment=1 # Center
    )

    story = []
    
    # Document Title and Banner
    story.append(Paragraph("EVENT TABULATION OFFICIAL REPORT", title_style))
    story.append(Paragraph(f"Exported metrics and academic standings for <b>{event_details['name']}</b>", subtitle_style))
    story.append(Spacer(1, 10))

    # Event Overview Grid (Key/Value Table)
    story.append(Paragraph("Event Specifications", h2_style))
    
    overview_data = [
        [
            Paragraph("<b>Event Name:</b>", body_bold), Paragraph(event_details['name'], body_regular),
            Paragraph("<b>Date:</b>", body_bold), Paragraph(event_details['date'], body_regular)
        ],
        [
            Paragraph("<b>Event Type:</b>", body_bold), Paragraph(event_details.get('event_type', 'N/A'), body_regular),
            Paragraph("<b>Time:</b>", body_bold), Paragraph(event_details['time'], body_regular)
        ],
        [
            Paragraph("<b>Category:</b>", body_bold), Paragraph(event_details['category'], body_regular),
            Paragraph("<b>Venue:</b>", body_bold), Paragraph(event_details['venue'], body_regular)
        ],
        [
            Paragraph("<b>Division:</b>", body_bold), Paragraph(event_details['division'], body_regular),
            Paragraph("<b>Tournament Type:</b>", body_bold), Paragraph(event_details.get('tournament_type', 'N/A'), body_regular)
        ],
        [
            Paragraph("<b>Department:</b>", body_bold), Paragraph(event_details['department'], body_regular),
            Paragraph("<b>Status:</b>", body_bold), Paragraph(event_details['status'], body_regular)
        ],
        [
            Paragraph("<b>Max Capacity:</b>", body_bold), Paragraph(str(event_details['max_participants']), body_regular),
            Paragraph("<b>Total Teams:</b>", body_bold), Paragraph(str(event_details['total_teams']), body_regular)
        ],
        [
            Paragraph("<b>Assigned Judges:</b>", body_bold), Paragraph(event_details.get('assigned_judges', 'None assigned'), body_regular),
            Paragraph("<b>Generated By:</b>", body_bold), Paragraph(str(event_info.get('generated_by', 'Admin')), body_regular)
        ]
    ]
    
    # 540 pt is the printable width ( letter is 612 - 72 )
    overview_table = Table(overview_data, colWidths=[100, 170, 80, 190])
    overview_table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, border_color),
        ('BACKGROUND', (0,0), (0,-1), stripe_color),
        ('BACKGROUND', (2,0), (2,-1), stripe_color),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('RIGHTPADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(overview_table)
    story.append(Spacer(1, 16))

    # Rankings and Scoring Table
    if 'rankings' in included_metrics or 'scores' in included_metrics:
        story.append(Paragraph("Rankings & Scorecards", h2_style))
        rank_headers = [
            Paragraph("<b>Rank</b>", body_bold),
            Paragraph("<b>Matchup</b>", body_bold),
            Paragraph("<b>Scores</b>", body_bold),
            Paragraph("<b>Winner</b>", body_bold),
            Paragraph("<b>Match #</b>", body_bold),
            Paragraph("<b>Status</b>", body_bold)
        ]
        
        rank_table_data = [rank_headers]
        for r in rankings:
            matchup_str = f"{r['team_a']} vs {r['team_b']}"
            scores_str = f"{r['score_a']} - {r['score_b']}"
            rank_table_data.append([
                Paragraph(str(r['rank']), body_center),
                Paragraph(matchup_str, body_regular),
                Paragraph(scores_str, body_center),
                Paragraph(r['winner'], body_regular),
                Paragraph(str(r['match_no']), body_center),
                Paragraph(r['status'], body_center)
            ])
            
        rank_table = Table(rank_table_data, colWidths=[40, 160, 80, 130, 50, 80])
        rank_styles = [
            ('BACKGROUND', (0,0), (-1,0), navy),
            ('GRID', (0,0), (-1,-1), 0.5, border_color),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
            ('RIGHTPADDING', (0,0), (-1,-1), 8),
        ]
        
        # Add white text style to all headers
        for col_idx in range(len(rank_headers)):
            rank_headers[col_idx].style.textColor = colors.white
            
        # Add alternate row backgrounds
        for idx in range(1, len(rank_table_data)):
            if idx % 2 == 0:
                rank_styles.append(('BACKGROUND', (0, idx), (-1, idx), stripe_color))
                
        rank_table.setStyle(TableStyle(rank_styles))
        story.append(rank_table)
        story.append(Spacer(1, 16))

    # Participants Directory
    if 'participants' in included_metrics:
        story.append(Paragraph("Participants Directory", h2_style))
        part_headers = [
            Paragraph("<b>Seed</b>", body_bold),
            Paragraph("<b>Team Name</b>", body_bold),
            Paragraph("<b>Department / Course</b>", body_bold),
            Paragraph("<b>Losses</b>", body_bold),
            Paragraph("<b>Members List</b>", body_bold)
        ]
        
        part_table_data = [part_headers]
        for p in participants:
            part_table_data.append([
                Paragraph(str(p['seed']), body_center),
                Paragraph(p['name'], body_regular),
                Paragraph(p['department'], body_regular),
                Paragraph(str(p['loss_count']), body_center),
                Paragraph(p['members'], body_regular)
            ])
            
        part_table = Table(part_table_data, colWidths=[40, 130, 130, 50, 190])
        part_styles = [
            ('BACKGROUND', (0,0), (-1,0), navy),
            ('GRID', (0,0), (-1,-1), 0.5, border_color),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
            ('RIGHTPADDING', (0,0), (-1,-1), 8),
        ]
        
        # Add white text to headers
        for col_idx in range(len(part_headers)):
            part_headers[col_idx].style.textColor = colors.white
            
        for idx in range(1, len(part_table_data)):
            if idx % 2 == 0:
                part_styles.append(('BACKGROUND', (0, idx), (-1, idx), stripe_color))
                
        part_table.setStyle(TableStyle(part_styles))
        story.append(part_table)
        story.append(Spacer(1, 16))

    # ── Helper: build a navy-header striped table ──
    def _styled_table(headers, data_rows, col_widths, center_idx=()):
        head_cells = [Paragraph(f"<b>{h}</b>", body_bold) for h in headers]
        for c in head_cells:
            c.style.textColor = colors.white
        table_data = [head_cells]
        for row in data_rows:
            cells = []
            for c_idx, val in enumerate(row):
                style = body_center if c_idx in center_idx else body_regular
                cells.append(Paragraph(str(val), style))
            table_data.append(cells)
        tbl = Table(table_data, colWidths=col_widths)
        tstyles = [
            ('BACKGROUND', (0, 0), (-1, 0), navy),
            ('GRID', (0, 0), (-1, -1), 0.5, border_color),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ]
        for idx in range(1, len(table_data)):
            if idx % 2 == 0:
                tstyles.append(('BACKGROUND', (0, idx), (-1, idx), stripe_color))
        tbl.setStyle(TableStyle(tstyles))
        return tbl

    # Bracket Overview Section
    if bracket_rounds:
        story.append(Paragraph("Bracket Overview", h2_style))
        for rnd in bracket_rounds:
            story.append(Paragraph(f"<b>{rnd['round']}</b>", body_bold))
            story.append(Spacer(1, 4))
            rows = [
                [m['match_no'], f"{m['team_a']} vs {m['team_b']}",
                 f"{m['score_a']} - {m['score_b']}", m['winner'], m['status']]
                for m in rnd['matches']
            ]
            story.append(_styled_table(
                ["Match #", "Matchup", "Score", "Winner", "Status"],
                rows, col_widths=[50, 190, 70, 130, 80], center_idx=(0, 2, 4),
            ))
            story.append(Spacer(1, 10))
        story.append(Paragraph(
            f"<b>Champion / Final Winner:</b> {final_winner}", body_bold))
        story.append(Spacer(1, 16))

    # Criteria-Based Results Section
    if criteria_results:
        story.append(Paragraph("Criteria-Based Results", h2_style))
        headers = ["Rank", "Candidate"] + criteria_names + ["Total"]
        n_crit = len(criteria_names)
        crit_w = (540 - 40 - 130 - 60) / n_crit if n_crit else 0
        col_widths = [40, 130] + [crit_w] * n_crit + [60]
        rows = []
        for r in criteria_results:
            row = [r['rank'], r['candidate']]
            row += [r['per_criterion'].get(cn, 0) for cn in criteria_names]
            row.append(r['total'])
            rows.append(row)
        center = tuple([0] + list(range(2, 2 + n_crit + 1)))
        story.append(_styled_table(headers, rows, col_widths, center_idx=center))
        story.append(Spacer(1, 16))

    # Leaderboard / Points Summary Section
    if leaderboard:
        story.append(Paragraph("Leaderboard / Points Summary", h2_style))
        rows = [[l['rank'], l['name'], l['group'], l['points'], l['detail']] for l in leaderboard]
        story.append(_styled_table(
            ["Rank", "Team / Participant", "Group", "Points", "Record / Detail"],
            rows, col_widths=[45, 160, 140, 60, 135], center_idx=(0, 3),
        ))
        story.append(Spacer(1, 16))

    # Remarks / Metadata Section
    remarks_text = event_info.get('remarks')
    if remarks_text:
        story.append(Paragraph("Remarks", h2_style))
        story.append(Paragraph(remarks_text, body_regular))
        story.append(Spacer(1, 16))

    # Audit Trail Section
    story.append(Paragraph("System Audit Trail", h2_style))
    audit_headers = [
        Paragraph("<b>Timestamp</b>", body_bold),
        Paragraph("<b>User</b>", body_bold),
        Paragraph("<b>Action</b>", body_bold),
        Paragraph("<b>Component</b>", body_bold),
        Paragraph("<b>Details</b>", body_bold)
    ]
    
    audit_table_data = [audit_headers]
    for l in audit_trail:
        audit_table_data.append([
            Paragraph(l['timestamp'], body_center),
            Paragraph(l['user'], body_regular),
            Paragraph(l['action_type'], body_center),
            Paragraph(l['repr'], body_regular),
            Paragraph(l['message'], body_regular)
        ])
        
    audit_table = Table(audit_table_data, colWidths=[90, 80, 60, 110, 200])
    audit_styles = [
        ('BACKGROUND', (0,0), (-1,0), navy),
        ('GRID', (0,0), (-1,-1), 0.5, border_color),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('RIGHTPADDING', (0,0), (-1,-1), 8),
    ]
    
    for col_idx in range(len(audit_headers)):
        audit_headers[col_idx].style.textColor = colors.white
        
    for idx in range(1, len(audit_table_data)):
        if idx % 2 == 0:
            audit_styles.append(('BACKGROUND', (0, idx), (-1, idx), stripe_color))
            
    audit_table.setStyle(TableStyle(audit_styles))
    story.append(KeepTogether([audit_table]))

    # Build the document using the NumberedCanvas to add dynamic page numbers
    doc.build(story, canvasmaker=NumberedCanvas)
    
    pdf_stream.seek(0)
    return pdf_stream
